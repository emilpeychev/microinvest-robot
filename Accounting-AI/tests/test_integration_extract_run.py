#!/usr/bin/env python3
"""Integration tests for extract.run end-to-end workbook generation."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except Exception:
    openpyxl = None
    OPENPYXL_AVAILABLE = False

from test_utils import load_module

extract = load_module("extract_invoices_v1_integration", "extract_invoices_v1.py")


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl is required for integration workbook test")
class ExtractRunIntegrationTests(unittest.TestCase):
    def test_run_generates_review_workbook_with_mandatory_flags(self):
        with tempfile.TemporaryDirectory() as tmp:
            base_dir = Path(tmp)
            client = "Client_IT"

            processed_dir = base_dir / "Clients" / client / "01_Processed"
            review_dir = base_dir / "Clients" / client / "02_Review"
            templates_dir = base_dir / "Templates"
            logs_dir = base_dir / "Logs"

            processed_dir.mkdir(parents=True)
            review_dir.mkdir(parents=True)
            templates_dir.mkdir(parents=True)
            logs_dir.mkdir(parents=True)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ExtractedInvoices"
            ws.append(
                [
                    "Client",
                    "File Name",
                    "Document Type (Invoice/Receipt)",
                    "Supplier/Customer",
                    "Invoice Number",
                    "Invoice Date",
                    "Net Amount",
                    "VAT Amount",
                    "Gross Amount",
                    "Currency",
                    "Confidence Score",
                    "Notes",
                    "Mandatory Review",
                ]
            )
            wb.save(templates_dir / "extracted_invoices.xlsx")

            (processed_dir / "Client_IT_2026-03-28_Invoice_Demo_120.00.pdf").write_bytes(b"%PDF-1.4\\n%EOF\\n")
            (processed_dir / "Client_IT_2026-03-28_Receipt_Photo_15.50.jpg").touch()
            (processed_dir / "Client_IT_2026-03-28_Bank_UniCredit_42.00.csv").touch()

            count = extract.run(base_dir=base_dir, client_name=client)
            self.assertEqual(count, 2)

            out_wb = openpyxl.load_workbook(review_dir / "extracted_invoices.xlsx")
            out_ws = out_wb.active
            headers = [c.value for c in out_ws[1]]
            col = {h: i for i, h in enumerate(headers)}

            rows = []
            for row in out_ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None and v != "" for v in row):
                    rows.append(row)

            self.assertEqual(len(rows), 2)
            by_file = {r[col["File Name"]]: r for r in rows}
            self.assertEqual(by_file["Client_IT_2026-03-28_Invoice_Demo_120.00.pdf"][col["Mandatory Review"]], "No")
            self.assertEqual(by_file["Client_IT_2026-03-28_Receipt_Photo_15.50.jpg"][col["Mandatory Review"]], "Yes")


if __name__ == "__main__":
    unittest.main()
